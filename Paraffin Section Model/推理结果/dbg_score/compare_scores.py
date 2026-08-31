"""Compare auto vs human-revised vs original manual scores."""
from pathlib import Path
import csv
from collections import Counter
import numpy as np
import openpyxl

base = Path(r"d:\BaiduNetdiskDownload\shibie")

def load_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    print(f"\n=== {path.name} ===")
    print(f"sheet={ws.title} max_row={ws.max_row} max_col={ws.max_column}")
    # print header
    for r in range(1, min(4, ws.max_row+1)):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
        print(f"  row{r}: {vals}")
    scores = {}
    avgs = {}
    for row in range(3, ws.max_row + 1):
        sample = ws.cell(row, 1).value
        if sample is None:
            continue
        sid = int(sample)
        views = {}
        for vid, col in zip((7, 8, 9), (2, 3, 4)):
            val = ws.cell(row, col).value
            if val is not None:
                try:
                    views[vid] = float(val)
                    scores[(sid, vid)] = float(val)
                except (TypeError, ValueError):
                    print(f"  skip {sid} col{col}={val!r}")
        avg = ws.cell(row, 5).value
        if avg is not None:
            try:
                avgs[sid] = float(avg)
            except (TypeError, ValueError):
                pass
        if not views:
            print(f"  empty sample {sid}")
    print(f"  n_cells={len(scores)} n_samples={len(avgs)}")
    if scores:
        arr = np.array(list(scores.values()))
        print(f"  min={arr.min():.2f} max={arr.max():.2f} mean={arr.mean():.2f} std={arr.std():.2f}")
        print(f"  dist={dict(sorted(Counter(int(round(v)) for v in scores.values()).items()))}")
    return scores, avgs

# find files
infer = base / "推理结果"
manual_dir = base / "人工标定"
revised = None
auto = None
orig = None
for p in infer.iterdir():
    if p.suffix.lower() != ".xlsx":
        continue
    name = p.name
    if "人工" in name or "修订" in name:
        revised = p
    elif "自动" in name:
        auto = p
    print("infer xlsx:", repr(p.name))
for p in manual_dir.iterdir():
    if p.suffix.lower() == ".xlsx":
        orig = p
        print("manual xlsx:", repr(p.name))

print("revised=", revised)
print("auto=", auto)
print("orig=", orig)

rev, rev_avg = load_xlsx(revised)
aut, aut_avg = load_xlsx(auto)
org, org_avg = load_xlsx(orig)

common = sorted(set(rev) & set(aut))
print(f"\n=== AUTO vs REVISED common={len(common)} ===")
diffs = []
rows = []
for k in common:
    a, r = aut[k], rev[k]
    d = a - r
    diffs.append(d)
    rows.append((abs(d), k[0], k[1], a, r, d))

diffs = np.array(diffs)
print(f"MAE={np.mean(np.abs(diffs)):.3f} RMSE={np.sqrt(np.mean(diffs**2)):.3f} bias(auto-rev)={np.mean(diffs):.3f}")
print(f"exact={np.mean(diffs==0):.1%}  within1={np.mean(np.abs(diffs)<=1):.1%}  within2={np.mean(np.abs(diffs)<=2):.1%}")
print(f"auto mean={np.mean([aut[k] for k in common]):.2f}  rev mean={np.mean([rev[k] for k in common]):.2f}")

# score_similarity as in metrics.py
pred = np.array([aut[k] for k in common])
gt = np.array([rev[k] for k in common])
mae = float(np.mean(np.abs(pred - gt)))
within1 = float(np.mean(np.abs(pred - gt) <= 1.0))
mae_sim = max(0.0, 1.0 - mae / 4.0)
sim = 0.6 * mae_sim + 0.4 * within1
print(f"score_similarity={sim:.1%}  mae_sim={mae_sim:.1%}  within1={within1:.1%}")

print("\n--- confusion auto(row) vs revised(col) ---")
auto_r = [int(round(aut[k])) for k in common]
rev_r = [int(round(rev[k])) for k in common]
for a_s in range(1, 11):
    line = []
    for r_s in range(1, 11):
        c = sum(1 for a, r in zip(auto_r, rev_r) if a == a_s and r == r_s)
        line.append(f"{c:3d}")
    print(f"auto{a_s:2d}: {' '.join(line)}")

print("\n--- largest |diff| ---")
for absd, sid, vid, a, r, d in sorted(rows, reverse=True)[:30]:
    print(f"  sample {sid:3d} view{vid}  auto={a:5.1f}  rev={r:5.1f}  diff={d:+5.1f}")

print("\n--- systematic: mean(auto-rev) by auto score ---")
from collections import defaultdict
by_auto = defaultdict(list)
by_rev = defaultdict(list)
for k in common:
    by_auto[int(round(aut[k]))].append(aut[k]-rev[k])
    by_rev[int(round(rev[k]))].append(aut[k]-rev[k])
for s in range(1, 11):
    if by_auto[s]:
        arr = np.array(by_auto[s])
        print(f"  auto={s}: n={len(arr):3d} mean_diff={arr.mean():+.2f} mae={np.abs(arr).mean():.2f}")
print("--- by revised ---")
for s in range(1, 11):
    if by_rev[s]:
        arr = np.array(by_rev[s])
        print(f"  rev={s}: n={len(arr):3d} mean_diff={arr.mean():+.2f} mae={np.abs(arr).mean():.2f}")

# vs original manual
print("\n=== ORIG vs REVISED ===")
c2 = sorted(set(rev) & set(org))
d2 = np.array([org[k]-rev[k] for k in c2])
print(f"common={len(c2)} MAE={np.mean(np.abs(d2)):.3f} exact={np.mean(d2==0):.1%} within1={np.mean(np.abs(d2)<=1):.1%}")
print(f"orig mean={np.mean([org[k] for k in c2]):.2f} rev mean={np.mean([rev[k] for k in c2]):.2f}")

print("\n=== ORIG vs AUTO ===")
c3 = sorted(set(aut) & set(org))
d3 = np.array([aut[k]-org[k] for k in c3])
print(f"common={len(c3)} MAE={np.mean(np.abs(d3)):.3f} exact={np.mean(d3==0):.1%} within1={np.mean(np.abs(d3)<=1):.1%}")

# also compare with csv features
csv_path = infer / "region3_scores.csv"
if csv_path.exists():
    import csv as csvmod
    feats = {}
    with csv_path.open(encoding="utf-8") as f:
        for row in csvmod.DictReader(f):
            k = (int(row["sample_id"]), int(row["view_id"]))
            feats[k] = row
    print("\n=== features vs revised (pearson) ===")
    keys = [k for k in common if k in feats]
    y = np.array([rev[k] for k in keys], dtype=float)
    for name in ["raw", "coverage", "gap_ratio", "break_count", "row_jitter"]:
        x = np.array([float(feats[k][name]) for k in keys])
        if x.std() > 0 and y.std() > 0:
            corr = np.corrcoef(x, y)[0,1]
        else:
            corr = float("nan")
        print(f"  {name:12s} corr={corr:+.3f}  mean={x.mean():.3f}")
    # MAE of CNN raw vs revised
    raw = np.array([float(feats[k]["raw"]) for k in keys])
    print(f"  CNN raw vs rev MAE={np.mean(np.abs(raw-y)):.3f}")

    print("\n--- features by revised score ---")
    for s in range(1, 11):
        ks = [k for k in keys if int(round(rev[k])) == s]
        if not ks:
            continue
        cov = np.mean([float(feats[k]["coverage"]) for k in ks])
        gap = np.mean([float(feats[k]["gap_ratio"]) for k in ks])
        brk = np.mean([float(feats[k]["break_count"]) for k in ks])
        jit = np.mean([float(feats[k]["row_jitter"]) for k in ks])
        auto_m = np.mean([aut[k] for k in ks])
        print(f"  rev={s:2d} n={len(ks):3d} auto_mean={auto_m:.2f} cov={cov:.3f} gap={gap:.3f} brk={brk:.2f} jit={jit:.1f}")
