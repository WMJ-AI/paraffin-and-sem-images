"""Fill 结果呈现表_新.xlsx without changing the old 结果呈现表.xlsx writer.

Old table keeps measured lengths / counts / areas and Python-computed averages.
New table writes raw measurements into 1/2/3 columns, then Excel formulas for:

- 相对树皮厚度 = (皮层长度 + 韧皮部长度) / 半径长度
- 相对韧皮部厚度 = 韧皮部长度 / 半径长度
- 相对木质部厚度 = 木质部长度 / 半径长度
- 导管密度 = 导管总数量 / 局部木质部面积，单位 vessels·mm⁻²
  （面积存的是 µm²，公式里乘 1e6 换成 mm²）
- 平均导管腔面积 = 导管（腔）总面积 / 导管总数量
- 导管腔面积占木质部比例 = 导管（腔）总面积 / 局部木质部面积
- 形成层 / 派生表的「平均」列 = AVERAGE(B:D)
- 汇总(均值) 引用各派生表平均列

局部木质部面积按新表 I9 说明：有韧皮部或髓时，用整图面积减去这些多余面积。
旧表仍是整图面积，本模块不改旧导出。
"""

from __future__ import annotations

from dataclasses import fields
from pathlib import Path

import openpyxl
from openpyxl.utils import quote_sheetname

from calibration.excel_export import (
    SampleRecord,
    empty_records,
    records_from_cambium_xlsx,
    records_from_region1_csv,
    records_from_region2_csv,
)

UM2_PER_MM2 = 1_000_000
VIEW_COLS = ("B", "C", "D")
MAX_SAMPLE = 63

TEMPLATE_NAME = "结果呈现表_新.xlsx"
OUTPUT_NAME = "结果呈现表_新_推理.xlsx"

RAW_SHEETS: dict[str, tuple[str, int]] = {
    "半径长度": ("radius", 2),
    "木质部长度": ("xylem", 2),
    "韧皮部长度": ("phloem", 2),
    "皮层长度": ("bark", 2),
    "导管总数量": ("vessel_count", 0),
    "导管（腔）总面积": ("vessel_area", 1),
    "局部韧皮部面积": ("phloem_area", 1),
    "局部髓面积": ("pith_area", 1),
}

CAMBIUM_SHEET = "形成层连续性评分1（坏）-10（好）"
XYLEM_AREA_SHEET = "局部木质部面积"
SUMMARY_SHEET = "汇总(均值)"

DERIVED_AVG_SHEETS = (
    "相对树皮厚度",
    "相对韧皮部厚度",
    "相对木质部厚度",
    "导管密度",
    "平均导管腔面积",
    "导管腔面积占木质部比例",
)

SUMMARY_AVG_REFS = (
    "相对树皮厚度",
    "相对韧皮部厚度",
    "相对木质部厚度",
    "导管密度",
    "平均导管腔面积",
    "导管腔面积占木质部比例",
    CAMBIUM_SHEET,
)


def _ref(sheet: str, col: str, row: int) -> str:
    return f"{quote_sheetname(sheet)}!{col}{row}"


def _div_formula(numer: str, denom: str) -> str:
    return f'IF(OR({denom}="",{denom}=0),"",({numer})/{denom})'


def _avg_formula(row: int) -> str:
    return f'IF(COUNT(B{row}:D{row})=0,"",AVERAGE(B{row}:D{row}))'


def _derived_view_formula(sheet: str, col: str, row: int) -> str:
    radius = _ref("半径长度", col, row)
    bark = _ref("皮层长度", col, row)
    phloem = _ref("韧皮部长度", col, row)
    xylem = _ref("木质部长度", col, row)
    count = _ref("导管总数量", col, row)
    lumen = _ref("导管（腔）总面积", col, row)
    xylem_area = _ref(XYLEM_AREA_SHEET, col, row)
    if sheet == "相对树皮厚度":
        return _div_formula(f"{bark}+{phloem}", radius)
    if sheet == "相对韧皮部厚度":
        return _div_formula(phloem, radius)
    if sheet == "相对木质部厚度":
        return _div_formula(xylem, radius)
    if sheet == "导管密度":
        return _div_formula(f"{count}*{UM2_PER_MM2}", xylem_area)
    if sheet == "平均导管腔面积":
        return _div_formula(lumen, count)
    if sheet == "导管腔面积占木质部比例":
        return _div_formula(lumen, xylem_area)
    raise ValueError(f"未知派生表: {sheet}")


def _xylem_area_for_new(rec: SampleRecord, idx: int) -> float | None:
    """New-table xylem area: full frame minus phloem/pith when either is present."""
    full = rec.xylem_area[idx]
    if full is None:
        return None
    phloem = rec.phloem_area[idx] or 0.0
    pith = rec.pith_area[idx] or 0.0
    if phloem == 0.0 and pith == 0.0:
        return float(full)
    return max(0.0, float(full) - float(phloem) - float(pith))


def _write_views(ws, sample_id: int, values: list[float | None], ndigits: int) -> None:
    row = sample_id + 1
    if ws.cell(row, 1).value in (None, ""):
        ws.cell(row, 1, sample_id)
    for i, val in enumerate(values, start=2):
        cell = ws.cell(row, i)
        if val is None:
            cell.value = None
        elif ndigits == 0:
            cell.value = int(round(val))
        else:
            cell.value = round(float(val), ndigits)


def _save_workbook(wb, path: Path) -> Path:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return path
    except PermissionError:
        alt = path.with_name(path.stem + "_已统计.xlsx")
        wb.save(alt)
        print(f"  {path.name} 正在被占用，已改写到 {alt.name}")
        return alt


def load_records_from_outputs(base_dir: Path) -> dict[int, SampleRecord]:
    records = empty_records(max_sample=MAX_SAMPLE)
    infer = base_dir / "推理结果"
    records = records_from_region1_csv(infer / "region1_metrics.csv", records)
    records = records_from_region2_csv(infer / "region2_metrics.csv", records)
    records = records_from_cambium_xlsx(infer / "形成层 评分_自动.xlsx", records)
    return records


def overlay_records(
    base_dir: Path, records: dict[int, SampleRecord] | None
) -> dict[int, SampleRecord]:
    merged = load_records_from_outputs(base_dir)
    if not records:
        return merged
    for sid, rec in records.items():
        if sid not in merged:
            merged[sid] = rec
            continue
        dest = merged[sid]
        for f in fields(SampleRecord):
            if f.name == "sample_id":
                continue
            src_vals = getattr(rec, f.name)
            dest_vals = getattr(dest, f.name)
            for i, val in enumerate(src_vals):
                if val is not None:
                    dest_vals[i] = val
    return merged


def fill_new_result_workbook(
    base_dir: Path,
    records: dict[int, SampleRecord] | None = None,
    template_path: Path | None = None,
    output_path: Path | None = None,
    also_fill_template: bool = True,
) -> Path | None:
    """Copy the new template, write raw values + Excel formulas, save filled workbooks."""
    template_path = template_path or (base_dir / TEMPLATE_NAME)
    output_path = output_path or (base_dir / OUTPUT_NAME)
    if not template_path.exists():
        print(f"[新结果表] 未找到模板 {template_path.name}，跳过")
        return None

    records = overlay_records(base_dir, records)
    wb = openpyxl.load_workbook(template_path)
    missing = [name for name in list(RAW_SHEETS) + list(DERIVED_AVG_SHEETS) + [XYLEM_AREA_SHEET, CAMBIUM_SHEET, SUMMARY_SHEET] if name not in wb.sheetnames]
    if missing:
        wb.close()
        raise KeyError(f"新表缺少工作表: {missing}")

    for title, (attr, ndigits) in RAW_SHEETS.items():
        ws = wb[title]
        for sid in range(1, MAX_SAMPLE + 1):
            rec = records[sid]
            _write_views(ws, sid, getattr(rec, attr), ndigits)

    ws_xy = wb[XYLEM_AREA_SHEET]
    for sid in range(1, MAX_SAMPLE + 1):
        rec = records[sid]
        areas = [_xylem_area_for_new(rec, i) for i in range(3)]
        _write_views(ws_xy, sid, areas, 1)

    ws_cam = wb[CAMBIUM_SHEET]
    for sid in range(1, MAX_SAMPLE + 1):
        rec = records[sid]
        _write_views(ws_cam, sid, rec.cambium_scores, 2)
        row = sid + 1
        ws_cam.cell(row, 5).value = f"={_avg_formula(row)}"

    for title in DERIVED_AVG_SHEETS:
        ws = wb[title]
        for sid in range(1, MAX_SAMPLE + 1):
            row = sid + 1
            if ws.cell(row, 1).value in (None, ""):
                ws.cell(row, 1, sid)
            for col in VIEW_COLS:
                ws[f"{col}{row}"] = f"={_derived_view_formula(title, col, row)}"
            ws.cell(row, 5).value = f"={_avg_formula(row)}"

    ws_sum = wb[SUMMARY_SHEET]
    for sid in range(1, MAX_SAMPLE + 1):
        row = sid + 1
        if ws_sum.cell(row, 1).value in (None, ""):
            ws_sum.cell(row, 1, sid)
        for col_idx, sheet in enumerate(SUMMARY_AVG_REFS, start=2):
            ws_sum.cell(row, col_idx).value = f"={_ref(sheet, 'E', row)}"

    saved = _save_workbook(wb, output_path)
    print(f"[新结果表] 已写入: {saved}")
    if also_fill_template and template_path.resolve() != saved.resolve():
        official = _save_workbook(wb, template_path)
        if official != saved:
            print(f"          模板同步: {official}")
    wb.close()
    return saved


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="用已有 CSV / 评分表填写 结果呈现表_新.xlsx")
    parser.add_argument("--base", type=Path, default=Path(__file__).resolve().parent.parent)
    parser.add_argument("--output", type=Path, default=None, help="默认 结果呈现表_新_推理.xlsx")
    parser.add_argument("--no-fill-template", action="store_true", help="只写推理副本，不改 结果呈现表_新.xlsx")
    args = parser.parse_args()
    fill_new_result_workbook(
        args.base,
        output_path=args.output,
        also_fill_template=not args.no_fill_template,
    )


if __name__ == "__main__":
    main()
