"""Cambium continuity scoring (1-10) calibrated against human-revised xlsx."""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

import cv2
import numpy as np

from calibration.region3.cambium import CambiumBox


@dataclass
class ContinuityFeatures:
    coverage: float
    gap_ratio: float
    break_count: int
    row_jitter: float


@dataclass
class CambiumScore:
    score: int
    raw: float
    features: ContinuityFeatures
    rule_hint: str


def extract_continuity_features(img: np.ndarray, box: CambiumBox) -> ContinuityFeatures:
    """Continuity features along the cambium band (absolute, not fixed-percentile)."""
    patch = img[box.y0 : box.y1, box.x0 : box.x1]
    if patch.size == 0:
        return ContinuityFeatures(0.0, 1.0, 99, 99.0)

    target_w = 800
    scale = max(1.0, patch.shape[1] / target_w)
    small = cv2.resize(patch, (target_w, max(24, int(patch.shape[0] / scale))), interpolation=cv2.INTER_AREA)
    gray = cv2.cvtColor(small, cv2.COLOR_BGR2GRAY).astype(np.float32)
    gray = cv2.GaussianBlur(gray, (5, 5), 0)
    h, w = gray.shape

    gy = cv2.Sobel(gray, cv2.CV_32F, 0, 1, ksize=3)
    row_e = np.mean(np.abs(gy), axis=1)
    ridge = int(np.argmax(row_e))
    band = np.abs(gy)[max(0, ridge - 4) : min(h, ridge + 5), :]
    strength = np.mean(band, axis=0)
    strength = np.convolve(strength, np.ones(11) / 11, mode="same")

    med = float(np.median(strength)) + 1e-6
    thr = 0.55 * med
    active = strength >= thr
    coverage = float(active.mean())
    gap_ratio = float(1.0 - coverage)

    break_count = 0
    inactive_run = 0
    min_gap = max(4, int(w * 0.012))
    for v in active:
        if v:
            if inactive_run >= min_gap:
                break_count += 1
            inactive_run = 0
        else:
            inactive_run += 1

    peaks = [int(np.argmax(gray[:, x])) for x in range(0, w, 2)]
    row_jitter = float(np.std(peaks)) if len(peaks) >= 5 else 20.0

    return ContinuityFeatures(
        coverage=coverage,
        gap_ratio=float(np.clip(gap_ratio, 0, 1)),
        break_count=int(break_count),
        row_jitter=row_jitter,
    )


def _composite(features: ContinuityFeatures) -> float:
    return (
        2.0 * features.coverage
        - 1.5 * features.gap_ratio
        - 0.35 * min(features.break_count, 12)
        - 0.01 * min(features.row_jitter, 60.0)
    )


def _rule_hint(score: int) -> str:
    """Hint must follow final 1-10 score (需求.docx rubric), not raw CV features."""
    if score <= 3:
        return "上下衔接的小区域已经出现断层 (低分)"
    if score <= 6:
        return "上下衔接的小区域排列散乱，即将断开 (中分)"
    if score <= 8:
        return "上下衔接较连续 (中高分)"
    return "上下衔接的小区域紧凑衔接 (高分)"


def score_from_features(features: ContinuityFeatures, calib: dict | None = None) -> CambiumScore:
    comp = _composite(features)
    if calib and calib.get("quantiles"):
        raw = _map_quantile(comp, calib["quantiles"])
    else:
        raw = 4.0 + comp * 1.8
    score = int(np.clip(round(raw), 1, 10))
    return CambiumScore(score=score, raw=float(raw), features=features, rule_hint=_rule_hint(score))


def score_cambium(img: np.ndarray, box: CambiumBox, calib: dict | None = None) -> CambiumScore:
    features = extract_continuity_features(img, box)
    return score_from_features(features, calib)


def score_cambium_cnn(img: np.ndarray, box: CambiumBox, score_model) -> CambiumScore:
    """Score with strip CNN trained on human-revised 1-10 labels; CV features kept for logging."""
    features = extract_continuity_features(img, box)
    crop = img[box.y0 : box.y1, box.x0 : box.x1]
    if crop.size == 0:
        return CambiumScore(score=1, raw=1.0, features=features, rule_hint=_rule_hint(1))
    raw = float(score_model.predict_crop_raw(crop))
    score = int(np.clip(round(raw), 1, 10))
    return CambiumScore(score=score, raw=raw, features=features, rule_hint=_rule_hint(score))


def resolve_score_xlsx(base_dir: Path) -> Path:
    """Prefer human-revised inference scores, else the original manual table."""
    infer = base_dir / "推理结果"
    if infer.exists():
        for path in sorted(infer.iterdir()):
            if path.suffix.lower() != ".xlsx":
                continue
            name = path.name
            if "修订" in name or "人工修订" in name:
                return path
    return base_dir / "人工标定" / "形成层 评分.xlsx"


def load_manual_scores(xlsx_path: Path) -> dict[tuple[int, int], float]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    scores: dict[tuple[int, int], float] = {}
    for row in range(3, ws.max_row + 1):
        sample = ws.cell(row, 1).value
        if sample is None:
            continue
        for view_id, col in zip((7, 8, 9), (2, 3, 4)):
            val = ws.cell(row, col).value
            if val is not None:
                scores[(int(sample), view_id)] = float(val)
    return scores


def _map_quantile(value: float, quantiles: list[list[float]]) -> float:
    if value <= quantiles[0][0]:
        return quantiles[0][1]
    if value >= quantiles[-1][0]:
        return quantiles[-1][1]
    for i in range(len(quantiles) - 1):
        x0, y0 = quantiles[i]
        x1, y1 = quantiles[i + 1]
        if x0 <= value <= x1:
            t = (value - x0) / (x1 - x0 + 1e-6)
            return y0 * (1 - t) + y1 * t
    return quantiles[len(quantiles) // 2][1]


def calibrate_weights(
    base_dir: Path,
    xlsx_path: Path | None = None,
    save_path: Path | None = None,
) -> dict:
    from calibration.io_util import imread, parse_name
    from calibration.region3.cambium import load_box_params, locate_cambium

    xlsx_path = xlsx_path or resolve_score_xlsx(base_dir)
    manual = load_manual_scores(xlsx_path)
    orig_dir = base_dir / "原图" / "第三区域"
    box_params = load_box_params(base_dir / "models" / "region3_box_params.json")

    pairs: list[tuple[float, float]] = []
    for path in sorted(orig_dir.glob("*.jpg")):
        meta = parse_name(path.name)
        if meta is None:
            continue
        key = (meta.sample_id, meta.view_id)
        if key not in manual:
            continue
        img = imread(path)
        if img is None:
            continue
        box = locate_cambium(img, box_params)
        if box is None:
            continue
        feat = extract_continuity_features(img, box)
        pairs.append((_composite(feat), manual[key]))

    if len(pairs) < 10:
        return {"quantiles": [[-2, 4], [0, 7], [2, 10]]}

    pairs.sort(key=lambda item: item[0])
    # Full rank mapping preserves manual score distribution on calibration set
    quantiles = [[comp, score] for comp, score in pairs]

    calib = {"quantiles": quantiles}

    if save_path:
        save_path.parent.mkdir(parents=True, exist_ok=True)
        save_path.write_text(json.dumps(calib, indent=2), encoding="utf-8")

    return calib


def load_weights(path: Path) -> dict:
    if not path.exists():
        return {"quantiles": [[-2, 4], [0, 7], [2, 10]]}
    return json.loads(path.read_text(encoding="utf-8"))
