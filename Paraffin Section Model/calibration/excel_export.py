"""Export inference results to 结果呈现表.xlsx."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl import Workbook


@dataclass
class SampleRecord:
    sample_id: int
    radius: list[float | None]
    xylem: list[float | None]
    phloem: list[float | None]
    bark: list[float | None]
    cambium_scores: list[float | None]
    vessel_count: list[float | None]
    vessel_area: list[float | None]
    xylem_area: list[float | None]
    phloem_area: list[float | None]
    pith_area: list[float | None]


def _avg(values: list[float | None]) -> float | None:
    nums = [v for v in values if v is not None]
    if not nums:
        return None
    return sum(nums) / len(nums)


def _round_avg(values: list[float | None], ndigits: int = 2) -> float | None:
    avg = _avg(values)
    if avg is None:
        return None
    return round(avg, ndigits)


def _write_detail_sheet(
    ws, sample_id: int, header: list[str], values: list[float | None], ndigits: int = 2
) -> None:
    row = sample_id + 1
    ws.cell(row, 1, sample_id)
    for i, val in enumerate(values, start=2):
        if val is None:
            ws.cell(row, i, None)
        elif ndigits == 0:
            ws.cell(row, i, int(round(val)))
        else:
            ws.cell(row, i, round(float(val), ndigits))
    avg = _round_avg(values, ndigits if ndigits > 0 else 2)
    if ndigits == 0 and avg is not None:
        avg = round(avg, 2)
    ws.cell(row, len(header), avg)


def export_results(
    template_path: Path,
    output_path: Path,
    records: dict[int, SampleRecord],
) -> Path:
    if template_path.exists():
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    sheets = {
        "radius": ("半径长度", ["样本序号", "半径长度1", "半径长度2", "半径长度3", "平均"]),
        "xylem": ("木质部长度", ["样本序号", "木质部长度1", "木质部长度2", "木质部长度3", "平均"]),
        "phloem": ("韧皮部长度", ["样本序号", "韧皮部长度1", "韧皮部长度2", "韧皮部长度3", "平均"]),
        "bark": ("树皮 皮层长度", ["样本序号", "树皮 皮层长度1", "树皮 皮层长度2", "树皮 皮层长度3", "平均"]),
        "vessel_count": ("导管总数量", ["样本序号", "导管总数量1", "导管总数量2", "导管总数量3", "平均"]),
        "vessel_area": ("导管（腔）总面积", ["样本序号", "导管（腔）总面积1", "导管（腔）总面积2", "导管（腔）总面积3", "平均"]),
        "xylem_area": ("局部木质部面积", ["样本序号", "局部木质部面积1", "局部木质部面积2", "局部木质部面积3", "平均"]),
        "phloem_area": ("局部韧皮部面积", ["样本序号", "局部韧皮部面积1", "局部韧皮部面积2", "局部韧皮部面积3", "平均"]),
        "pith_area": ("局部髓面积", ["样本序号", "局部髓面积1", "局部髓面积2", "局部髓面积3", "平均"]),
        "cambium": (
            "形成层连续性评分1（坏）-10（好）",
            ["样本序号", "形成层连续性评分1", "形成层连续性评分2", "形成层连续性评分3", "平均"],
        ),
    }

    def _sheet(title: str, header: list[str]):
        if title in wb.sheetnames:
            return wb[title]
        # tolerate 树皮/皮层 vs 树皮 皮层 naming
        for name in wb.sheetnames:
            if name.replace("/", " ").replace("  ", " ").strip() == title.replace("/", " ").strip():
                return wb[name]
        ws = wb.create_sheet(title)
        for col, h in enumerate(header, start=1):
            ws.cell(1, col, h)
        return ws

    for key, (title, header) in sheets.items():
        _sheet(title, header)

    summary_name = "汇总(均值)"
    if summary_name not in wb.sheetnames:
        ws_sum = wb.create_sheet(summary_name, 0)
        headers = [
            "样本序号",
            "半径长度",
            "木质部长度",
            "韧皮部长度",
            "树皮/皮层长度",
            "导管总数量",
            "导管（腔）总面积",
            "局部木质部面积",
            "形成层连续性评分1（坏）-10（好）",
        ]
        for col, name in enumerate(headers, start=1):
            ws_sum.cell(1, col, name)
    else:
        ws_sum = wb[summary_name]

    for sample_id in sorted(records):
        rec = records[sample_id]
        _write_detail_sheet(_sheet(*sheets["radius"]), sample_id, sheets["radius"][1], rec.radius, ndigits=2)
        _write_detail_sheet(_sheet(*sheets["xylem"]), sample_id, sheets["xylem"][1], rec.xylem, ndigits=2)
        _write_detail_sheet(_sheet(*sheets["phloem"]), sample_id, sheets["phloem"][1], rec.phloem, ndigits=2)
        _write_detail_sheet(_sheet(*sheets["bark"]), sample_id, sheets["bark"][1], rec.bark, ndigits=2)
        _write_detail_sheet(
            _sheet(*sheets["vessel_count"]), sample_id, sheets["vessel_count"][1], rec.vessel_count, ndigits=0
        )
        _write_detail_sheet(
            _sheet(*sheets["vessel_area"]), sample_id, sheets["vessel_area"][1], rec.vessel_area, ndigits=1
        )
        _write_detail_sheet(
            _sheet(*sheets["xylem_area"]), sample_id, sheets["xylem_area"][1], rec.xylem_area, ndigits=1
        )
        _write_detail_sheet(
            _sheet(*sheets["phloem_area"]), sample_id, sheets["phloem_area"][1], rec.phloem_area, ndigits=1
        )
        _write_detail_sheet(
            _sheet(*sheets["pith_area"]), sample_id, sheets["pith_area"][1], rec.pith_area, ndigits=1
        )
        _write_detail_sheet(
            _sheet(*sheets["cambium"]), sample_id, sheets["cambium"][1], rec.cambium_scores, ndigits=2
        )

        row = sample_id + 1
        ws_sum.cell(row, 1, sample_id)
        ws_sum.cell(row, 2, _round_avg(rec.radius, 2))
        ws_sum.cell(row, 3, _round_avg(rec.xylem, 2))
        ws_sum.cell(row, 4, _round_avg(rec.phloem, 2))
        ws_sum.cell(row, 5, _round_avg(rec.bark, 2))
        ws_sum.cell(row, 6, _round_avg(rec.vessel_count, 2))
        ws_sum.cell(row, 7, _round_avg(rec.vessel_area, 1))
        ws_sum.cell(row, 8, _round_avg(rec.xylem_area, 1))
        ws_sum.cell(row, 9, _round_avg(rec.cambium_scores, 2))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    return _save_workbook(wb, output_path)


def _save_workbook(wb, path: Path) -> Path:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return path
    except PermissionError:
        alt = path.with_name(path.stem + "_已统计.xlsx")
        wb.save(alt)
        print(f"  {path.name} 正在被占用，已改写到 {alt.name}")
        return alt


def _ensure_sheet(wb, title: str, header: list[str]):
    if title in wb.sheetnames:
        return wb[title]
    ws = wb.create_sheet(title)
    for col, name in enumerate(header, start=1):
        ws.cell(1, col, name)
    return ws


def write_region2_excel(
    template_path: Path,
    output_path: Path,
    records: dict[int, SampleRecord],
    detail_rows: list[dict] | None = None,
) -> Path:
    """Fill 导管总数量 / 导管（腔）总面积 / 局部木质部面积 and 汇总 columns 6–8.

    Does not overwrite region-1 lengths or region-3 scores if the output workbook
    already exists.
    """
    src = output_path if output_path.exists() else template_path
    if src.exists():
        wb = openpyxl.load_workbook(src)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    count_header = ["样本序号", "导管总数量1", "导管总数量2", "导管总数量3", "平均"]
    area_header = ["样本序号", "导管（腔）总面积1", "导管（腔）总面积2", "导管（腔）总面积3", "平均"]
    xylem_header = ["样本序号", "局部木质部面积1", "局部木质部面积2", "局部木质部面积3", "平均"]
    phloem_header = ["样本序号", "局部韧皮部面积1", "局部韧皮部面积2", "局部韧皮部面积3", "平均"]
    pith_header = ["样本序号", "局部髓面积1", "局部髓面积2", "局部髓面积3", "平均"]
    ws_n = _ensure_sheet(wb, "导管总数量", count_header)
    ws_a = _ensure_sheet(wb, "导管（腔）总面积", area_header)
    ws_x = _ensure_sheet(wb, "局部木质部面积", xylem_header)
    ws_ph = _ensure_sheet(wb, "局部韧皮部面积", phloem_header)
    ws_pi = _ensure_sheet(wb, "局部髓面积", pith_header)

    summary_name = "汇总(均值)"
    if summary_name not in wb.sheetnames:
        ws_sum = wb.create_sheet(summary_name, 0)
        headers = [
            "样本序号",
            "半径长度",
            "木质部长度",
            "韧皮部长度",
            "树皮/皮层长度",
            "导管总数量",
            "导管（腔）总面积",
            "局部木质部面积",
            "形成层连续性评分1（坏）-10（好）",
        ]
        for col, name in enumerate(headers, start=1):
            ws_sum.cell(1, col, name)
    else:
        ws_sum = wb[summary_name]

    for sample_id in sorted(records):
        rec = records[sample_id]
        counts = [None if v is None else int(round(v)) for v in rec.vessel_count]
        areas = [None if v is None else round(v, 1) for v in rec.vessel_area]
        xylems = [None if v is None else round(v, 1) for v in rec.xylem_area]
        phloems = [None if v is None else round(v, 1) for v in rec.phloem_area]
        piths = [None if v is None else round(v, 1) for v in rec.pith_area]
        _write_detail_sheet(ws_n, sample_id, count_header, counts, ndigits=2)
        _write_detail_sheet(ws_a, sample_id, area_header, areas, ndigits=1)
        _write_detail_sheet(ws_x, sample_id, xylem_header, xylems, ndigits=1)
        _write_detail_sheet(ws_ph, sample_id, phloem_header, phloems, ndigits=1)
        _write_detail_sheet(ws_pi, sample_id, pith_header, piths, ndigits=1)
        row = sample_id + 1
        ws_sum.cell(row, 1, sample_id)
        ws_sum.cell(row, 6, _round_avg(counts, 2))
        ws_sum.cell(row, 7, _round_avg(areas, 1))
        ws_sum.cell(row, 8, _round_avg(xylems, 1))

    note_title = "第二区域说明"
    if note_title in wb.sheetnames:
        del wb[note_title]
    ws_n2 = wb.create_sheet(note_title)
    notes = [
        "第二区域统计（需求：导管总数量、导管（腔）总面积、局部木质部面积）",
        "面积单位：µm²。由右下角 100 µm 黑标尺换算（像素面积 × (µm/px)²）。",
        "导管总数量：木质部内等效直径≥16 µm（腔面积≥200 µm²）的导管个数，贴近人工黄框里实际腔的大小。",
        "导管（腔）总面积：全部导管腔面积之和（仅木质部内，不含韧皮部/髓）。",
        "局部韧皮部面积、局部髓面积：只出现在图片左右两侧一小部分；没有则为 0。",
        "局部木质部面积 = 整幅视野面积（与人工标定橙色外框一致，即整图面积）。",
        "每一样品 3 张图（视野 4/5/6）分别填入明细 1/2/3，平均写入「汇总(均值)」第 6–8 列。",
    ]
    for i, text in enumerate(notes, start=1):
        ws_n2.cell(i, 1, text)

    if detail_rows:
        title = "第二区域明细"
        header = [
            "文件名",
            "样本序号",
            "视野(4/5/6)",
            "导管总数量",
            "导管（腔）总面积_um2",
            "局部木质部面积_um2",
            "局部韧皮部面积_um2",
            "局部髓面积_um2",
            "um_per_pixel",
        ]
        if title in wb.sheetnames:
            del wb[title]
        ws_d = wb.create_sheet(title)
        for col, name in enumerate(header, start=1):
            ws_d.cell(1, col, name)
        ws_d.cell(
            1,
            11,
            "面积单位: µm²。局部木质部面积=整幅视野面积（与人工橙框一致）。导管仅统计木质部内腔体。",
        )
        for i, row in enumerate(detail_rows, start=2):
            ws_d.cell(i, 1, row.get("filename"))
            ws_d.cell(i, 2, int(row.get("sample_id", 0) or 0))
            ws_d.cell(i, 3, int(row.get("view_id", 0) or 0))
            ws_d.cell(i, 4, int(float(row.get("vessel_count", 0) or 0)))
            ws_d.cell(i, 5, round(float(row.get("lumen_area_um2", 0) or 0), 1))
            ws_d.cell(i, 6, round(float(row.get("xylem_area_um2", 0) or 0), 1))
            ws_d.cell(i, 7, round(float(row.get("phloem_area_um2", 0) or 0), 1))
            ws_d.cell(i, 8, round(float(row.get("pith_area_um2", 0) or 0), 1))
            ws_d.cell(i, 9, float(row.get("um_per_pixel", 0) or 0))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    return _save_workbook(wb, output_path)


def records_from_region1_csv(csv_path: Path, records: dict[int, SampleRecord] | None = None) -> dict[int, SampleRecord]:
    import csv

    if records is None:
        records = empty_records()
    if not csv_path.exists():
        return records
    with csv_path.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            sid = int(row["sample_id"])
            vid = int(row["view_id"])
            idx = view_index(vid, "第一区域")
            if idx is None or sid not in records:
                continue
            rec = records[sid]

            def _f(key: str) -> float | None:
                v = row.get(key)
                if v is None or v == "":
                    return None
                return float(v)

            rec.radius[idx] = _f("radius_um")
            rec.xylem[idx] = _f("xylem_um")
            rec.phloem[idx] = _f("phloem_um")
            rec.bark[idx] = _f("bark_um")
    return records


def records_from_cambium_xlsx(
    xlsx_path: Path, records: dict[int, SampleRecord] | None = None
) -> dict[int, SampleRecord]:
    """Load 形成层 评分_自动.xlsx (样品编号 + 视野1/2/3)."""
    if records is None:
        records = empty_records()
    if not xlsx_path.exists():
        return records
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    for row in range(3, ws.max_row + 1):
        sid = ws.cell(row, 1).value
        if sid is None:
            continue
        try:
            sid = int(sid)
        except (TypeError, ValueError):
            continue
        if sid not in records:
            continue
        scores: list[float | None] = []
        for col in (2, 3, 4):
            v = ws.cell(row, col).value
            if v is None or v == "":
                scores.append(None)
            else:
                scores.append(float(v))
        records[sid].cambium_scores = scores
    wb.close()
    return records


def records_from_region2_csv(csv_path: Path, records: dict[int, SampleRecord] | None = None) -> dict[int, SampleRecord]:
    import csv

    if records is None:
        records = empty_records()
    if not csv_path.exists():
        return records
    with csv_path.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            sid = int(row["sample_id"])
            vid = int(row["view_id"])
            idx = view_index(vid, "第二区域")
            if idx is None or sid not in records:
                continue
            rec = records[sid]
            rec.vessel_count[idx] = float(row["vessel_count"])
            rec.vessel_area[idx] = float(row["lumen_area_um2"])
            rec.xylem_area[idx] = float(row["xylem_area_um2"])
            rec.phloem_area[idx] = float(row.get("phloem_area_um2") or 0)
            rec.pith_area[idx] = float(row.get("pith_area_um2") or 0)
    return records


def empty_records(max_sample: int = 63) -> dict[int, SampleRecord]:
    out: dict[int, SampleRecord] = {}
    for sid in range(1, max_sample + 1):
        out[sid] = SampleRecord(
            sample_id=sid,
            radius=[None, None, None],
            xylem=[None, None, None],
            phloem=[None, None, None],
            bark=[None, None, None],
            cambium_scores=[None, None, None],
            vessel_count=[None, None, None],
            vessel_area=[None, None, None],
            xylem_area=[None, None, None],
            phloem_area=[None, None, None],
            pith_area=[None, None, None],
        )
    return out


def view_index(view_id: int, region: str) -> int | None:
    if region == "第一区域" and view_id in (1, 2, 3):
        return view_id - 1
    if region == "第二区域" and view_id in (4, 5, 6):
        return view_id - 4
    if region == "第三区域" and view_id in (7, 8, 9):
        return view_id - 7
    return None
