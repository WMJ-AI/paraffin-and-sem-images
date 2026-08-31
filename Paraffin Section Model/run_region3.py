"""Region-3 pipeline: mid-rule box + calibrate to >=85% IoU + scoring."""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path

from calibration.excel_export import SampleRecord, empty_records, view_index
from calibration.io_util import imread, imwrite, list_region_images, parse_name
from calibration.merge import stitch_horizontal
from calibration.region3.box_calibrate import optimize_box_params
from calibration.region3.cambium import draw_cambium_annotation, load_box_params, locate_cambium
from calibration.region3.scoring import (
    calibrate_weights,
    load_weights,
    score_cambium,
    score_cambium_cnn,
)


def _write_score_xlsx(path: Path, by_sample: dict[int, dict[int, int]]) -> None:
    """Write scores in the same layout as 人工标定/形成层 评分.xlsx."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "样品编号"
    ws["B1"] = "形成层评分（从1-到10）"
    ws["B2"] = "视野1"
    ws["C2"] = "视野2"
    ws["D2"] = "视野3"
    ws["E2"] = "平均"

    row = 3
    for sample_id in sorted(by_sample):
        views = by_sample[sample_id]
        v1 = views.get(7)
        v2 = views.get(8)
        v3 = views.get(9)
        vals = [v for v in (v1, v2, v3) if v is not None]
        avg = round(sum(vals) / len(vals), 2) if vals else None
        ws.cell(row, 1, sample_id)
        ws.cell(row, 2, v1)
        ws.cell(row, 3, v2)
        ws.cell(row, 4, v3)
        ws.cell(row, 5, avg)
        row += 1

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_score_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "filename",
        "sample_id",
        "view_id",
        "score",
        "raw",
        "coverage",
        "gap_ratio",
        "break_count",
        "row_jitter",
        "rule_hint",
        "backend",
    ]
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def run_region3(
    base_dir: Path,
    calibrate: bool = False,
    optimize_box: bool = True,
    target_iou: float = 0.85,
    score_only: bool = False,
    use_cnn: bool = True,
    records: dict[int, SampleRecord] | None = None,
) -> dict[int, SampleRecord]:
    region = "第三区域"
    orig_dir = base_dir / "原图" / region
    manual_dir = base_dir / "人工标定" / region
    auto_dir = base_dir / "自动标定" / region
    merge_dir = base_dir / "合并标定" / region
    weights_path = base_dir / "models" / "region3_score_weights.json"
    box_params_path = base_dir / "models" / "region3_box_params.json"
    cnn_path = base_dir / "models" / "yolo" / "score_regressor.pt"
    scores_xlsx = base_dir / "推理结果" / "形成层 评分_自动.xlsx"
    scores_csv = base_dir / "推理结果" / "region3_scores.csv"

    auto_dir.mkdir(parents=True, exist_ok=True)
    merge_dir.mkdir(parents=True, exist_ok=True)
    weights_path.parent.mkdir(parents=True, exist_ok=True)
    scores_xlsx.parent.mkdir(parents=True, exist_ok=True)

    manual_names = {p.name for p in manual_dir.glob("*.jpg")}
    for extra in merge_dir.glob("*.jpg"):
        if extra.name not in manual_names:
            extra.unlink()

    if records is None:
        records = empty_records()

    if score_only:
        optimize_box = False

    # 1) Mid-rule box params: calibrate / iterate vs manual until IoU >= target
    if optimize_box or (calibrate and not score_only) or not box_params_path.exists():
        print(f"[第三区域] 规则卡中间 + 算法校准框 (目标 mean IoU>={target_iou:.0%})...")
        box_params, box_stats = optimize_box_params(
            base_dir, target=target_iou, save_path=box_params_path
        )
        print(
            f"  框准确率 mean_iou={box_stats['mean_iou']:.3f} "
            f"IoU>=0.5占比={box_stats['acc_iou50']:.3f} "
            f">=人工尺寸占比={box_stats['ge_size_rate']:.3f}"
        )
        if box_stats["mean_iou"] < target_iou:
            print(f"  警告: 未达 {target_iou:.0%}，已保存当前最优规则，可继续调参迭代")
    else:
        box_params = load_box_params(box_params_path)
        print(f"[第三区域] 加载框规则(不改框): {box_params_path}")

    score_model = None
    calib = None
    backend = "cv"
    if use_cnn and cnn_path.exists():
        from ml.score_model import ScoreModel

        score_model = ScoreModel(cnn_path)
        backend = "cnn"
        print(f"[第三区域] 使用 CNN 评分: {cnn_path}")
    elif calibrate or not weights_path.exists():
        print("[第三区域] 使用人工修订/人工评分表校准打分权重...")
        calib = calibrate_weights(base_dir, save_path=weights_path)
        print(f"  分位映射: {len(calib.get('quantiles', []))} 锚点")
        backend = "cv"
    else:
        calib = load_weights(weights_path)
        print(f"[第三区域] 加载评分权重: {weights_path}")

    files = list_region_images(base_dir, region, view_ids=(7, 8, 9))
    mode = "仅重打分(框不变)" if score_only else "标定+评分"
    print(f"[第三区域] {mode} {len(files)} 张图 (backend={backend})...")

    by_sample: dict[int, dict[int, int]] = defaultdict(dict)
    csv_rows: list[dict] = []
    for i, orig_path in enumerate(files):
        meta = parse_name(orig_path.name)
        if meta is None:
            continue

        img = imread(orig_path)
        if img is None:
            continue

        manual_path = manual_dir / orig_path.name
        manual = imread(manual_path) if manual_path.exists() else None

        band = locate_cambium(img, box_params)
        if band is None:
            print(f"  跳过(未定位形成层): {orig_path.name}")
            continue

        if score_model is not None:
            result = score_cambium_cnn(img, band, score_model)
        else:
            result = score_cambium(img, band, calib)
        annotated = draw_cambium_annotation(img, band, score=result.score)

        imwrite(auto_dir / orig_path.name, annotated)

        if manual is not None:
            imwrite(merge_dir / orig_path.name, stitch_horizontal(manual, annotated))

        idx = view_index(meta.view_id, region)
        if idx is not None and meta.sample_id in records:
            records[meta.sample_id].cambium_scores[idx] = float(result.score)

        by_sample[meta.sample_id][meta.view_id] = int(result.score)
        feat = result.features
        csv_rows.append(
            {
                "filename": orig_path.name,
                "sample_id": meta.sample_id,
                "view_id": meta.view_id,
                "score": result.score,
                "raw": round(result.raw, 4),
                "coverage": round(feat.coverage, 4),
                "gap_ratio": round(feat.gap_ratio, 4),
                "break_count": feat.break_count,
                "row_jitter": round(feat.row_jitter, 4),
                "rule_hint": result.rule_hint,
                "backend": backend,
            }
        )

        if (i + 1) % 20 == 0 or i == 0:
            print(f"  {i + 1}/{len(files)}: {orig_path.name} score={result.score} {result.rule_hint}")

    if by_sample:
        _write_score_xlsx(scores_xlsx, by_sample)
        _write_score_csv(scores_csv, csv_rows)
        print(f"[第三区域] 分数表(同人工格式): {scores_xlsx}")
        print(f"[第三区域] 明细 CSV: {scores_csv}")

    print(f"[第三区域] 完成 -> {auto_dir}")
    print(f"          合并标定 -> {merge_dir} (左:人工标定 右:自动标定)")
    return records


def main() -> None:
    parser = argparse.ArgumentParser(description="第三区域：形成层空心框标定与连续性评分")
    parser.add_argument("--base", type=Path, default=Path(__file__).resolve().parent)
    parser.add_argument("--calibrate", action="store_true", help="重新校准评分权重 + 框规则")
    parser.add_argument("--no-optimize-box", action="store_true", help="跳过框规则优化，直接加载已有参数")
    parser.add_argument(
        "--score-only",
        action="store_true",
        help="只微调评分：不改框规则/标注几何，仅用现有框重打分并更新分数叠加",
    )
    parser.add_argument("--no-cnn", action="store_true", help="禁用 CNN，回退到 CV 特征分位映射")
    parser.add_argument("--retrain-score", action="store_true", help="用人工修订表重新训练评分 CNN 后再打分")
    parser.add_argument("--target-iou", type=float, default=0.85, help="框 mean IoU 目标")
    args = parser.parse_args()
    if args.retrain_score:
        from ml.train_score import train_score

        train_score(args.base)
    records = run_region3(
        args.base,
        calibrate=args.calibrate,
        optimize_box=not args.no_optimize_box,
        target_iou=args.target_iou,
        score_only=args.score_only,
        use_cnn=not args.no_cnn,
    )
    from calibration.excel_export_new import fill_new_result_workbook

    fill_new_result_workbook(args.base, records=records)


if __name__ == "__main__":
    main()
