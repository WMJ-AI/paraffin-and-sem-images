"""Excel writer + golden comparison helper."""
from __future__ import annotations

from pathlib import Path

import numpy as np
import openpyxl
from openpyxl import Workbook

from pair_tvw import VesselPair
from segment import Vessel


def summarize(vessels: list[Vessel], pairs: list[VesselPair]) -> dict:
    ais = np.array([v.ai_um2 for v in vessels], dtype=float)
    dis = np.array([v.di_um for v in vessels], dtype=float)
    # Primary TVW = mean of 3–5 wall normals (复查计划); median kept on pair object
    tvws = np.array([p.tvw_mean_um for p in pairs], dtype=float) if pairs else np.array([])
    cwrs = np.array([p.cwr_pair for p in pairs], dtype=float) if pairs else np.array([])
    dh = float(np.sum(dis**5) / np.sum(dis**4)) if len(dis) else float("nan")
    n_line = sum(len(p.lines) for p in pairs)

    def desc(x: np.ndarray) -> dict:
        if len(x) == 0:
            return {k: None for k in ("n", "mean", "sd", "median", "q1", "q3", "min", "max", "cv")}
        return {
            "n": int(len(x)),
            "mean": float(np.mean(x)),
            "sd": float(np.std(x, ddof=1)) if len(x) > 1 else 0.0,
            "median": float(np.median(x)),
            "q1": float(np.percentile(x, 25)),
            "q3": float(np.percentile(x, 75)),
            "min": float(np.min(x)),
            "max": float(np.max(x)),
            "cv": float(np.std(x, ddof=1) / np.mean(x) * 100) if len(x) > 1 and np.mean(x) else 0.0,
        }

    return {
        "n_vessel": len(vessels),
        "n_pair": len(pairs),
        "n_line": n_line,
        "Ai": desc(ais),
        "Di": desc(dis),
        "TVW": desc(tvws),
        "CWR": desc(cwrs),
        "Dh": dh,
        "Amean": float(np.mean(ais)) if len(ais) else None,
        "Amax": float(np.max(ais)) if len(ais) else None,
        "Amin": float(np.min(ais)) if len(ais) else None,
        "Dmean": float(np.mean(dis)) if len(dis) else None,
        "Dmax": float(np.max(dis)) if len(dis) else None,
        "Dmin": float(np.min(dis)) if len(dis) else None,
        "TVW_mean": float(np.mean(tvws)) if len(tvws) else None,
        "CWR_mean": float(np.mean(cwrs)) if len(cwrs) else None,
    }


def write_sample_excel(
    out_path: Path,
    sample_name: str,
    vessels: list[Vessel],
    pairs: list[VesselPair],
    um_per_px: float,
    scale_method: str,
) -> dict:
    summary = summarize(vessels, pairs)
    wb = Workbook()

    # --- 汇总 ---
    ws = wb.active
    ws.title = f"{sample_name}汇总"
    ws["A1"] = f"{sample_name}扫描电镜指标汇总（自动管线）"
    ws["A2"] = (
        f"A/D/Dh：{summary['n_vessel']}个合格导管；CWR：{summary['n_pair']}个相贴导管对；"
        f"标尺={um_per_px:.6f} μm/px（{scale_method}）"
    )
    ws.append([])
    ws.append(["统计层级", "数值", "单位", "解释"])
    ws.append(["n_vessel", summary["n_vessel"], "个导管", "A/D/Dh技术观测"])
    ws.append(["n_pair", summary["n_pair"], "个导管对", "CWR主统计单位"])
    ws.append(["n_line", summary["n_line"], "条线", "壁内技术测量"])
    ws.append([])
    ws.append(["统计量", "Ai (μm²)", "Di (μm)", "TVW_pair (μm)", "CWR_pair"])
    for label, key in [
        ("n", "n"),
        ("均值", "mean"),
        ("SD", "sd"),
        ("中位数", "median"),
        ("Q1", "q1"),
        ("Q3", "q3"),
        ("最小", "min"),
        ("最大", "max"),
        ("CV (%)", "cv"),
    ]:
        ws.append(
            [
                label,
                summary["Ai"][key],
                summary["Di"][key],
                summary["TVW"][key],
                summary["CWR"][key],
            ]
        )
    ws.append([])
    ws.append(["Dh", summary["Dh"], "μm", "ΣDi⁵/ΣDi⁴"])

    # --- 导管母表 ---
    ws2 = wb.create_sheet("导管母表")
    ws2.append(
        [
            "序号",
            "图像文件",
            "ROI编号",
            "状态",
            "Ai (μm²)",
            "Di (μm)",
            "周长 (μm)",
            "中心X (px)",
            "中心Y (px)",
            "Pair_ID",
            "来源/质控",
        ]
    )
    for i, v in enumerate(vessels, start=1):
        ws2.append(
            [
                i,
                v.image_name,
                v.vessel_id,
                v.status,
                v.ai_um2,
                v.di_um,
                v.peri_um,
                round(v.cx),
                round(v.cy),
                v.pair_id,
                v.reason,
            ]
        )

    # --- 导管对 ---
    ws3 = wb.create_sheet("导管对_CWR")
    ws3.append(
        [
            "Pair_ID",
            "图像文件",
            "导管1",
            "Ai1",
            "Di1",
            "导管2",
            "Ai2",
            "Di2",
            "n_line",
            "TVW中位数",
            "TVW均值",
            "TVW_SD",
            "CWR1",
            "CWR2",
            "CWR_pair",
        ]
    )
    for p in pairs:
        ws3.append(
            [
                p.pair_id,
                p.image_name,
                p.v1.vessel_id,
                p.v1.ai_um2,
                p.v1.di_um,
                p.v2.vessel_id,
                p.v2.ai_um2,
                p.v2.di_um,
                len(p.lines),
                p.tvw_median_um,
                p.tvw_mean_um,
                p.tvw_sd_um,
                p.cwr1,
                p.cwr2,
                p.cwr_pair,
            ]
        )

    # --- TVW线 ---
    ws4 = wb.create_sheet("TVW线测量明细")
    ws4.append(
        [
            "Pair_ID",
            "图像文件",
            "Line_ID",
            "起点X",
            "起点Y",
            "终点X",
            "终点Y",
            "长度(px)",
            "标定",
            "TVW(μm)",
            "分位",
        ]
    )
    for p in pairs:
        for ln in p.lines:
            ws4.append(
                [
                    ln.pair_id,
                    ln.image_name,
                    ln.line_id,
                    ln.x0,
                    ln.y0,
                    ln.x1,
                    ln.y1,
                    ln.length_px,
                    ln.um_per_px,
                    ln.tvw_um,
                    ln.quantile,
                ]
            )

    # --- Sheet1 统计结果 ---
    ws5 = wb.create_sheet("Sheet1 统计结果")
    ws5.append(
        [
            "样品序号",
            "导管腔面积 Ai (μm2)（可记录多个）",
            "最大导管腔面积 Amax (μm2)",
            "最小导管腔面积 Amin (μm2)",
            "平均导管腔面积 Amean (μm2)",
            "等效圆直径 Di (μm)（可记录多个）",
            "最大等效圆直径 Dmax (μm)",
            "最小等效圆直径 Dmin (μm)",
            "平均等效圆直径 Dmean (μm)",
            "水力加权直径 Dh(μm)",
            "双导管壁厚 TVW (μm)",
            "壁强化指数 CWR",
            "说明",
        ]
    )
    ws5.append(
        [
            sample_name,
            f"见导管母表（{summary['n_vessel']}个）",
            summary["Amax"],
            summary["Amin"],
            summary["Amean"],
            f"见导管母表（{summary['n_vessel']}个）",
            summary["Dmax"],
            summary["Dmin"],
            summary["Dmean"],
            summary["Dh"],
            summary["TVW_mean"],
            summary["CWR_mean"],
            f"自动管线；标尺{um_per_px:.6f}μm/px；CWR用{summary['n_pair']}个pair",
        ]
    )
    for i in range(2, 64):
        ws5.append([f"样品{i}"])

    # Placeholder explanation sheets (short pointers to Wand template)
    ws6 = wb.create_sheet("Sheet1 黄色指标样貌说明")
    ws6["A1"] = "黄色区域指标说明同 Wand 模板（Ai/Amax/Amin/Amean/TVW）。本自动结果仅统计绿色合格导管，红/黄/白过滤。"
    ws7 = wb.create_sheet("Sheet1 绿色指标计算公式")
    ws7["A1"] = (
        "Di=2√(Ai/π)；Dh=ΣDi^5/ΣDi^4；CWR_pair=avg[(TVW/Di1)^2,(TVW/Di2)^2]；"
        "TVW主值=壁内3–5条局部法线的均值（中位数另存对照）。Ai=不规则闭合轮廓面积。"
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return summary


def load_golden(xlsx: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    # 读取外部黄金 Excel 的既有工作表名（磁盘模板字段，非本仓库规则命名）
    ws = wb["样品1汇总"]
    golden = {
        "n_vessel": ws["B5"].value,
        "n_pair": ws["B6"].value,
        "n_line": ws["B8"].value,
        "Amean": ws["B13"].value,
        "Dmean": ws["C13"].value,
        "TVW_mean": ws["D13"].value,
        "CWR_mean": ws["E13"].value,
        "Dh": ws["B25"].value,
        "Amax": ws["B19"].value,
        "Amin": ws["B18"].value,
    }
    vessels = []
    ws2 = wb["导管母表_31"]
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
        if row[0] is None:
            continue
        vessels.append(
            {
                "image": row[1],
                "id": row[2],
                "ai": row[5],
                "di": row[6],
                "cx": row[8],
                "cy": row[9],
                "pair": row[14],
            }
        )
    pairs = []
    ws3 = wb["导管对_CWR_9"]
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, values_only=True):
        if row[0] is None:
            continue
        pairs.append(
            {
                "id": row[0],
                "image": row[1],
                "v1": row[2],
                "v2": row[5],
                "tvw": row[10],
                "cwr": row[18],
                "cx1": None,
                "cy1": None,
            }
        )
    # attach centers
    id_to_xy = {v["id"]: (v["cx"], v["cy"], v["image"]) for v in vessels}
    for p in pairs:
        if p["v1"] in id_to_xy:
            p["cx1"], p["cy1"], _ = id_to_xy[p["v1"]]
        if p["v2"] in id_to_xy:
            p["cx2"], p["cy2"], _ = id_to_xy[p["v2"]]
    golden["vessels"] = vessels
    golden["pairs"] = pairs
    return golden


def match_vessels(auto: list[Vessel], golden_vessels: list[dict], max_dist: float = 40.0):
    """Greedy nearest-center matching."""
    remaining = set(range(len(golden_vessels)))
    matches = []
    unmatched_auto = []
    for v in auto:
        best_i, best_d = None, 1e9
        for i in remaining:
            g = golden_vessels[i]
            if g["image"] != v.image_name:
                continue
            d = (v.cx - g["cx"]) ** 2 + (v.cy - g["cy"]) ** 2
            if d < best_d:
                best_d = d
                best_i = i
        if best_i is not None and best_d <= max_dist**2:
            remaining.remove(best_i)
            g = golden_vessels[best_i]
            matches.append((v, g, float(np.sqrt(best_d))))
        else:
            unmatched_auto.append(v)
    unmatched_golden = [golden_vessels[i] for i in sorted(remaining)]
    return matches, unmatched_auto, unmatched_golden


def compare_report(auto_summary: dict, golden: dict, matches, unmatched_auto, unmatched_golden) -> str:
    lines = []
    lines.append("=== 自动 vs 人工黄金值 ===")
    keys = [
        ("n_vessel", "导管数"),
        ("n_pair", "导管对数"),
        ("Amean", "Amean"),
        ("Dmean", "Dmean"),
        ("Dh", "Dh"),
        ("TVW_mean", "TVW均值"),
        ("CWR_mean", "CWR均值"),
        ("Amax", "Amax"),
        ("Amin", "Amin"),
    ]
    for k, name in keys:
        a = auto_summary.get(k)
        g = golden.get(k)
        if a is None or g is None:
            lines.append(f"{name}: auto={a} golden={g}")
            continue
        if isinstance(a, (int, np.integer)) or k.startswith("n_"):
            lines.append(f"{name}: auto={a} golden={g}  diff={a-g}")
        else:
            pct = (a - g) / g * 100 if g else float("nan")
            lines.append(f"{name}: auto={a:.4f} golden={g:.4f}  diff%={pct:.2f}%")

    lines.append("")
    lines.append(
        f"导管匹配: matched={len(matches)}  auto_only={len(unmatched_auto)}  golden_only={len(unmatched_golden)}"
    )
    if matches:
        ai_err = []
        for v, g, d in matches:
            if v.ai_um2 and g["ai"]:
                ai_err.append(abs(v.ai_um2 - g["ai"]) / g["ai"] * 100)
        lines.append(
            f"匹配导管 Ai 相对误差: mean={np.mean(ai_err):.2f}%  median={np.median(ai_err):.2f}%  max={np.max(ai_err):.2f}%"
        )
    if unmatched_golden:
        lines.append("漏检黄金导管:")
        for g in unmatched_golden:
            lines.append(f"  {g['id']} @({g['cx']},{g['cy']}) Ai={g['ai']:.1f}")
    if unmatched_auto:
        lines.append("多检自动导管:")
        for v in unmatched_auto:
            lines.append(f"  {v.vessel_id} @({v.cx:.0f},{v.cy:.0f}) Ai={v.ai_um2:.1f}")
    return "\n".join(lines)
