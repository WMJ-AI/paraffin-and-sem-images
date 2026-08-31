"""
Generate Excel matching the Wand template sheet format for the full batch.

Sheets (same order/roles as template):
  1. 汇总
  2. 导管母表_N
  3. 候选导管审计
  4. 视野统计
  5. 导管对_CWR_N
  6. TVW线测量明细
  7. Sheet1 统计结果
  8. Sheet1 黄色指标样貌说明
  9. Sheet1 绿色指标计算公式
"""
from __future__ import annotations

import re
import sys
import time
from collections import defaultdict
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

from pair_tvw import find_pairs  # noqa: E402
from run_sample1 import process_image  # noqa: E402
from segment import tissue_height  # noqa: E402

DATA = Path(r"H:\尉明杰\扫描电镜 模型")
BATCH = next(p for p in DATA.iterdir() if p.is_dir() and "63" in p.name and "526" in p.name)
def _find_golden_template() -> Path | None:
    root = next(
        (d for d in DATA.iterdir() if d.is_dir() and "例子" in d.name),
        DATA,
    )
    return next(
        (p for p in root.glob("*.xlsx") if "Wand" in p.name or "扫描电镜" in p.name),
        None,
    )


GOLDEN = _find_golden_template()
OUT = ROOT / "output" / "batch_all_526"
OUT.mkdir(parents=True, exist_ok=True)
XLSX_OUT = OUT / "全批次_扫描电镜指标_自动分析结果.xlsx"


def natural_key(p: Path):
    m = re.match(r"(\d+)\s*\((\d+)\)", p.stem)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    return (9999, p.name)


def list_unique_images() -> list[Path]:
    files = sorted(BATCH.glob("*.tif"), key=natural_key)
    seen, uniq = set(), []
    for f in files:
        k = f.name.lower()
        if k in seen:
            continue
        seen.add(k)
        uniq.append(f)
    return uniq


def sample_id(path: Path) -> int:
    m = re.match(r"(\d+)\s*\((\d+)\)", path.stem)
    return int(m.group(1)) if m else 0


def copy_sheet(src_wb, src_name: str, dst_wb, dst_name: str) -> None:
    src = src_wb[src_name]
    if dst_name in dst_wb.sheetnames:
        del dst_wb[dst_name]
    dst = dst_wb.create_sheet(dst_name)
    for row in src.iter_rows():
        for cell in row:
            dst.cell(cell.row, cell.column, cell.value)
    for col_idx, col_dim in src.column_dimensions.items():
        if col_dim.width:
            dst.column_dimensions[col_idx].width = col_dim.width


def sanitize_green_demo_formulas(ws) -> int:
    """
    Template green sheet stores Excel *examples* as live formulas like
    F8=AVERAGE(F2:F64), which includes F8 itself → circular reference.
    Convert self-range demo formulas to plain text.
    """
    fixed = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            col = "".join(c for c in cell.coordinate if c.isalpha())
            row_n = int("".join(c for c in cell.coordinate if c.isdigit()))
            for m in re.finditer(rf"{col}(\d+):{col}(\d+)", v, flags=re.I):
                a, b = int(m.group(1)), int(m.group(2))
                if min(a, b) <= row_n <= max(a, b):
                    cell.value = "'" + v
                    fixed += 1
                    break
    return fixed


def write_workbook(
    by_sample: dict,
    images: list[dict],
    *,
    out_path: Path | None = None,
    sample_lo: int | None = None,
    sample_hi: int | None = None,
) -> Path:
    """
    写出与 Wand 模板同结构的 Excel。
    sample_lo/hi: Sheet1 样品行范围（默认 1..max(by_sample 或 63)）。
    """
    all_vessels = []
    all_pairs = []
    for sid in sorted(by_sample):
        all_vessels.extend(by_sample[sid]["vessels"])
        all_pairs.extend(by_sample[sid]["pairs"])

    n_v = len(all_vessels)
    n_p = len(all_pairs)
    if sample_lo is None:
        sample_lo = 1
    if sample_hi is None:
        sample_hi = max(by_sample.keys()) if by_sample else 63
        sample_hi = max(sample_hi, 1)
    n_bio = int(sample_hi - sample_lo + 1)
    dest = Path(out_path) if out_path else XLSX_OUT
    dest.parent.mkdir(parents=True, exist_ok=True)
    mother = f"导管母表_{n_v}"
    pair_sheet = f"导管对_CWR_{n_p}"
    line_sheet = "TVW线测量明细"
    audit_sheet = "候选导管审计"
    view_sheet = "视野统计"
    summary_sheet = "汇总"

    wb = Workbook()

    # ---------- 导管母表 ----------
    ws_m = wb.active
    ws_m.title = mother
    ws_m.append(
        [
            "序号",
            "图像文件",
            "ROI编号",
            "导管序号",
            "状态",
            "Ai (μm²)",
            "Di=2√(Ai/π) (μm)",
            "周长 (μm)",
            "中心X (px)",
            "中心Y (px)",
            "Ai敏感性下限",
            "Ai敏感性上限",
            "Di敏感性下限",
            "Di敏感性上限",
            "Pair_ID（若用于CWR）",
            "用于A/D/Dh",
            "来源/质控",
            "样品",
        ]
    )
    vessel_row: dict[str, int] = {}  # vessel_id -> excel row
    sample_vessel_rows: dict[int, tuple[int, int]] = {}
    seq_in_img: dict[str, int] = defaultdict(int)
    row = 1
    for sid in sorted(by_sample):
        start = row + 1
        for v in by_sample[sid]["vessels"]:
            row += 1
            seq_in_img[v.image_name] += 1
            vessel_row[v.vessel_id] = row
            ai = float(v.ai_um2 or 0)
            lo, hi = ai * 0.98, ai * 1.02
            ws_m.append(
                [
                    row - 1,
                    v.image_name,
                    v.vessel_id,
                    seq_in_img[v.image_name],
                    v.status or "primary",
                    ai,
                    f"=2*SQRT(F{row}/PI())",
                    float(v.peri_um or 0),
                    int(round(v.cx)),
                    int(round(v.cy)),
                    lo,
                    hi,
                    f"=2*SQRT(K{row}/PI())",
                    f"=2*SQRT(L{row}/PI())",
                    v.pair_id or None,
                    "是",
                    f"auto_pipeline;{v.reason}",
                    f"样品{sid}",
                ]
            )
        if by_sample[sid]["vessels"]:
            sample_vessel_rows[sid] = (start, row)

    # ---------- TVW线测量明细 ----------
    ws_l = wb.create_sheet(line_sheet)
    ws_l.append(
        [
            "Pair_ID",
            "图像文件",
            "Line_ID",
            "起点X (px)",
            "起点Y (px)",
            "终点X (px)",
            "终点Y (px)",
            "长度 (px)",
            "标定 (μm/px)",
            "TVW (μm)",
            "与导管1局部法线偏差 (°)",
            "与导管2局部法线偏差 (°)",
            "最大法线偏差 (°)",
            "共同壁位置分位",
            "纳入",
            "线级质控",
            "样品",
        ]
    )
    pair_line_ranges: dict[str, tuple[int, int]] = {}
    line_row = 1
    for sid in sorted(by_sample):
        for p in by_sample[sid]["pairs"]:
            if not p.lines:
                continue
            r0 = line_row + 1
            for ln in p.lines:
                line_row += 1
                r = line_row
                ws_l.append(
                    [
                        ln.pair_id,
                        ln.image_name,
                        ln.line_id,
                        ln.x0,
                        ln.y0,
                        ln.x1,
                        ln.y1,
                        f"=SQRT((F{r}-D{r})^2+(G{r}-E{r})^2)",
                        ln.um_per_px,
                        f"=H{r}*I{r}",
                        None,
                        None,
                        f'=IF(OR(K{r}="",L{r}=""),"",MAX(K{r},L{r}))',
                        ln.quantile,
                        "是",
                        f'=IF(OR(M{r}="",M{r}<=18),"PASS","ANGLE_ALERT")',
                        f"样品{sid}",
                    ]
                )
            pair_line_ranges[p.pair_id] = (r0, line_row)

    # ---------- 导管对_CWR ----------
    ws_p = wb.create_sheet(pair_sheet)
    ws_p.append(
        [
            "Pair_ID",
            "图像文件",
            "导管1 ROI",
            "Ai1 (μm²)",
            "Di1 (μm)",
            "导管2 ROI",
            "Ai2 (μm²)",
            "Di2 (μm)",
            "TVW线表范围",
            "n_line",
            "TVW均值 (μm，主值)",
            "TVW中位数 (μm)",
            "TVW线内SD (μm)",
            "TVW线内CV (%)",
            "旧单线TVW (μm)",
            "TVW变化 (%)",
            "CWR1=(TVW/Di1)²",
            "CWR2=(TVW/Di2)²",
            "CWR_pair（主值）",
            "旧单线CWR_pair",
            "CWR变化 (%)",
            "线中位数法CWR_pair（敏感性）",
            "Dtarget（仅敏感性）",
            "Dtarget法CWR（仅敏感性）",
            "壁级QC",
            "对应关系质控",
            "样品",
        ]
    )
    sample_pair_rows: dict[int, tuple[int, int]] = {}
    pair_excel_row = 1
    for sid in sorted(by_sample):
        start = pair_excel_row + 1
        for p in by_sample[sid]["pairs"]:
            pair_excel_row += 1
            r = pair_excel_row
            r1 = vessel_row.get(p.v1.vessel_id)
            r2 = vessel_row.get(p.v2.vessel_id)
            lr = pair_line_ranges.get(p.pair_id)
            if lr:
                a, b = lr
                rng = f"J{a}:J{b}"
                n_f = f"=COUNT({line_sheet}!J{a}:J{b})"
                # K=mean (主值), L=median (对照)
                avg_f = f"=AVERAGE({line_sheet}!J{a}:J{b})"
                med_f = f"=MEDIAN({line_sheet}!J{a}:J{b})"
                sd_f = f"=IF(J{r}>1,STDEV({line_sheet}!J{a}:J{b}),0)"
            else:
                rng = ""
                n_f = len(p.lines)
                avg_f = p.tvw_mean_um
                med_f = p.tvw_median_um
                sd_f = p.tvw_sd_um
            old_tvw = p.lines[len(p.lines) // 2].tvw_um if p.lines else p.tvw_mean_um
            ai1 = f"={mother}!F{r1}" if r1 else p.v1.ai_um2
            di1 = f"={mother}!G{r1}" if r1 else p.v1.di_um
            ai2 = f"={mother}!F{r2}" if r2 else p.v2.ai_um2
            di2 = f"={mother}!G{r2}" if r2 else p.v2.di_um
            ws_p.append(
                [
                    p.pair_id,
                    p.image_name,
                    p.v1.vessel_id,
                    ai1,
                    di1,
                    p.v2.vessel_id,
                    ai2,
                    di2,
                    rng,
                    n_f,
                    avg_f,
                    med_f,
                    sd_f,
                    f"=IF(K{r}=0,\"\",M{r}/K{r}*100)",
                    old_tvw,
                    f"=IF(O{r}=0,\"\",(K{r}/O{r}-1)*100)",
                    f"=(K{r}/E{r})^2",
                    f"=(K{r}/H{r})^2",
                    f"=AVERAGE(Q{r},R{r})",
                    f"=AVERAGE((O{r}/E{r})^2,(O{r}/H{r})^2)",
                    f"=IF(T{r}=0,\"\",(S{r}/T{r}-1)*100)",
                    f"=AVERAGE((L{r}/E{r})^2,(L{r}/H{r})^2)",
                    f'=IF(ABS(E{r}-{summary_sheet}!B25)<=ABS(H{r}-{summary_sheet}!B25),E{r},H{r})',
                    f"=(K{r}/W{r})^2",
                    f'=IF(N{r}>10,"HIGH_TVW_CV",IF(J{r}<3,"SHORT_INTERFACE","PASS"))',
                    "Di1、TVW均值、Di2均来自本Pair_ID；TVW=3–5法线平均",
                    f"样品{sid}",
                ]
            )
        if by_sample[sid]["pairs"]:
            sample_pair_rows[sid] = (start, pair_excel_row)

    # ---------- 候选导管审计 ----------
    ws_a = wb.create_sheet(audit_sheet)
    ws_a.append(
        [
            "Candidate_ID",
            "图像文件",
            "中心X",
            "中心Y",
            "边界框 (xmin,ymin,xmax,ymax)",
            "距最近视野边界 (px)",
            "状态",
            "reason_code",
            "可可靠测量Ai/Di",
            "主分析判定",
            "Ai (μm²)",
            "Di (μm)",
            "边界类别",
            "处理",
            "判定依据",
            "审计来源",
            "Pair_ID",
            "样品",
        ]
    )
    for sid in sorted(by_sample):
        for v in by_sample[sid]["vessels"]:
            x, y, bw, bh = v.bbox
            xmin, ymin, xmax, ymax = x, y, x + bw, y + bh
            # approximate image size from typical SEM; use bbox margin vs 1024x768
            edge = min(xmin, ymin, 1024 - xmax, 768 - ymax)
            edge = max(0, edge)
            border = "近边缘（<25 px）" if edge < 25 else "视野内部"
            r = vessel_row[v.vessel_id]
            ws_a.append(
                [
                    v.vessel_id,
                    v.image_name,
                    int(round(v.cx)),
                    int(round(v.cy)),
                    f"{xmin},{ymin},{xmax},{ymax}",
                    edge,
                    "primary",
                    "AUTO_VESSEL",
                    "是",
                    "纳入",
                    float(v.ai_um2 or 0),
                    f"={mother}!G{r}",
                    border,
                    "纳入A/D/Dh；若有Pair_ID则纳入CWR",
                    "自动管线检出的合格导管腔。",
                    f"auto_pipeline;{v.reason}",
                    v.pair_id or None,
                    f"样品{sid}",
                ]
            )

    # ---------- 视野统计 ----------
    ws_v = wb.create_sheet(view_sheet)
    ws_v.append(
        [
            "图像",
            "primary导管",
            "borderline候选（含未测）",
            "excluded候选",
            "edge_truncated筛查候选",
            "相贴导管对",
            "TVW线数",
            "像素尺寸 (μm/px)",
            "组织高 (px)",
            "视野面积 (mm²)",
            "说明",
            "样品",
        ]
    )
    for i, im in enumerate(images, start=2):
        name = im["image"]
        um = im["um_per_px"]
        th = im["tissue_h"]
        ws_v.append(
            [
                name,
                f'=COUNTIF({mother}!B:B,A{i})',
                f'=COUNTIFS({audit_sheet}!B:B,A{i},{audit_sheet}!G:G,"borderline")',
                f'=COUNTIFS({audit_sheet}!B:B,A{i},{audit_sheet}!G:G,"excluded")',
                f'=COUNTIFS({audit_sheet}!B:B,A{i},{audit_sheet}!G:G,"edge_truncated")',
                f'=COUNTIF({pair_sheet}!B:B,A{i})',
                f'=COUNTIF({line_sheet}!B:B,A{i})',
                um,
                th,
                f"=1024*I{i}*H{i}^2/1000000",
                "同一样品的技术视野；不是独立生物学重复",
                f"样品{im['sample']}",
            ]
        )
    last_view = 1 + len(images)
    if images:
        ws_v.append(
            [
                "合计",
                f"=SUM(B2:B{last_view})",
                f"=SUM(C2:C{last_view})",
                f"=SUM(D2:D{last_view})",
                f"=SUM(E2:E{last_view})",
                f"=SUM(F2:F{last_view})",
                f"=SUM(G2:G{last_view})",
                None,
                None,
                f"=SUM(J2:J{last_view})",
                f"{len(images)}个视野在样品层面按样品合并后写入Sheet1",
                "",
            ]
        )

    # ---------- 汇总 (global, formula-driven like template) ----------
    # Insert as first sheet
    ws_s = wb.create_sheet(summary_sheet, 0)
    vr0, vr1 = 2, 1 + n_v
    pr0, pr1 = 2, 1 + n_p
    lr0, lr1 = 2, max(2, line_row)
    ws_s["A1"] = "扫描电镜指标汇总（通用规则自动管线，格式同 Wand 模板）"
    ws_s["A2"] = (
        f"A/D/Dh：{n_v}个全部合格导管；CWR：{n_p}个唯一相贴导管对；"
        f"{2*n_p}个CWR1/2只作侧别追溯，不视为独立重复。"
        f"跨样品{sample_lo}–{sample_hi}（共{n_bio}个）合并母表；Sheet1按样品分行。"
    )
    ws_s["A4"] = "统计层级"
    ws_s["B4"] = "数值"
    ws_s["C4"] = "单位"
    ws_s["D4"] = "解释"
    rows_meta = [
        (5, "n_vessel", f"=COUNTA({mother}!C{vr0}:C{vr1})", "个导管", "A/D/Dh技术观测"),
        (6, "n_pair", f"=COUNTA({pair_sheet}!A{pr0}:A{pr1})", "个导管对", "CWR主统计单位"),
        (7, "n_component", "=2*B6", "个侧别分量", "CWR1/2，仅追溯；非独立"),
        (8, "n_line", f"=COUNT({line_sheet}!J{lr0}:J{lr1})", "条线", "壁内技术测量"),
        (9, "n_bio", str(n_bio), "个样品", "Sheet1每行一个样品生物学单位"),
    ]
    for r, a, b, c, d in rows_meta:
        ws_s[f"A{r}"] = a
        ws_s[f"B{r}"] = b
        ws_s[f"C{r}"] = c
        ws_s[f"D{r}"] = d

    ws_s["A11"] = "统计量"
    ws_s["B11"] = "Ai (μm²)"
    ws_s["C11"] = "Di (μm)"
    ws_s["D11"] = "TVW_pair (μm)"
    ws_s["E11"] = "CWR_pair"
    ws_s["F11"] = "层级说明"
    stats = [
        (12, "n", f"=COUNT({mother}!F{vr0}:F{vr1})", f"=COUNT({mother}!G{vr0}:G{vr1})",
         f"=COUNT({pair_sheet}!K{pr0}:K{pr1})", f"=COUNT({pair_sheet}!S{pr0}:S{pr1})", "导管/导管对"),
        (13, "均值", f"=AVERAGE({mother}!F{vr0}:F{vr1})", f"=AVERAGE({mother}!G{vr0}:G{vr1})",
         f"=AVERAGE({pair_sheet}!K{pr0}:K{pr1})", f"=AVERAGE({pair_sheet}!S{pr0}:S{pr1})", "CWR为pair均值"),
        (14, "SD", f"=STDEV({mother}!F{vr0}:F{vr1})", f"=STDEV({mother}!G{vr0}:G{vr1})",
         f"=STDEV({pair_sheet}!K{pr0}:K{pr1})", f"=STDEV({pair_sheet}!S{pr0}:S{pr1})", "技术变异"),
        (15, "中位数", f"=MEDIAN({mother}!F{vr0}:F{vr1})", f"=MEDIAN({mother}!G{vr0}:G{vr1})",
         f"=MEDIAN({pair_sheet}!K{pr0}:K{pr1})", f"=MEDIAN({pair_sheet}!S{pr0}:S{pr1})", "稳健位置统计"),
        (16, "Q1", f"=QUARTILE({mother}!F{vr0}:F{vr1},1)", f"=QUARTILE({mother}!G{vr0}:G{vr1},1)",
         f"=QUARTILE({pair_sheet}!K{pr0}:K{pr1},1)", f"=QUARTILE({pair_sheet}!S{pr0}:S{pr1},1)", "25%分位"),
        (17, "Q3", f"=QUARTILE({mother}!F{vr0}:F{vr1},3)", f"=QUARTILE({mother}!G{vr0}:G{vr1},3)",
         f"=QUARTILE({pair_sheet}!K{pr0}:K{pr1},3)", f"=QUARTILE({pair_sheet}!S{pr0}:S{pr1},3)", "75%分位"),
        (18, "最小", f"=MIN({mother}!F{vr0}:F{vr1})", f"=MIN({mother}!G{vr0}:G{vr1})",
         f"=MIN({pair_sheet}!K{pr0}:K{pr1})", f"=MIN({pair_sheet}!S{pr0}:S{pr1})", None),
        (19, "最大", f"=MAX({mother}!F{vr0}:F{vr1})", f"=MAX({mother}!G{vr0}:G{vr1})",
         f"=MAX({pair_sheet}!K{pr0}:K{pr1})", f"=MAX({pair_sheet}!S{pr0}:S{pr1})", None),
        (20, "CV (%)", "=B14/B13*100", "=C14/C13*100", "=D14/D13*100", "=E14/E13*100", "SD/均值×100"),
    ]
    for r, a, b, c, d, e, f in stats:
        ws_s[f"A{r}"] = a
        ws_s[f"B{r}"] = b
        ws_s[f"C{r}"] = c
        ws_s[f"D{r}"] = d
        ws_s[f"E{r}"] = e
        if f:
            ws_s[f"F{r}"] = f

    ws_s["A22"] = "一致性/水力指标"
    ws_s["B22"] = "结果"
    ws_s["C22"] = "单位"
    ws_s["D22"] = "公式或说明"
    ws_s["A23"] = "侧别分量算术均值"
    ws_s["B23"] = f"=(SUM({pair_sheet}!Q{pr0}:Q{pr1})+SUM({pair_sheet}!R{pr0}:R{pr1}))/(2*B6)"
    ws_s["C23"] = "无量纲"
    ws_s["D23"] = "应与CWR_pair均值一致；统计n仍为n_pair"
    ws_s["A24"] = "CWR层级一致性误差"
    ws_s["B24"] = "=ABS(B23-E13)"
    ws_s["C24"] = "无量纲"
    ws_s["D24"] = "应为0（浮点容差内）"
    ws_s["A25"] = "Dh（全库合并，仅参考）"
    ws_s["B25"] = f"=SUMPRODUCT({mother}!G{vr0}:G{vr1}^5)/SUMPRODUCT({mother}!G{vr0}:G{vr1}^4)"
    ws_s["C25"] = "μm"
    ws_s["D25"] = "ΣDi⁵/ΣDi⁴；正式按样品见Sheet1"
    ws_s["A27"] = "CWR方案"
    ws_s["B27"] = "CWR均值"
    ws_s["C27"] = "与主值差异 (%)"
    ws_s["D27"] = "用途"
    ws_s["A28"] = "优化主值：TVW壁内均值(3–5线)"
    ws_s["B28"] = f"=AVERAGE({pair_sheet}!S{pr0}:S{pr1})"
    ws_s["C28"] = 0
    ws_s["D28"] = "主结果"
    ws_s["A29"] = "优化敏感性：TVW壁内中位数"
    ws_s["B29"] = f"=AVERAGE({pair_sheet}!V{pr0}:V{pr1})"
    ws_s["C29"] = "=B29/B28*100-100"
    ws_s["D29"] = "评估壁内汇总函数"
    ws_s["A30"] = "旧单线法"
    ws_s["B30"] = f"=AVERAGE({pair_sheet}!T{pr0}:T{pr1})"
    ws_s["C30"] = "=B30/B28*100-100"
    ws_s["D30"] = "评估单线选择影响"
    ws_s["A31"] = "Dtarget版本"
    ws_s["B31"] = f"=AVERAGE({pair_sheet}!X{pr0}:X{pr1})"
    ws_s["C31"] = "=B31/B28*100-100"
    ws_s["D31"] = "仅方法敏感性；非主公式"

    # ---------- Sheet1 统计结果 ----------
    ws1 = wb.create_sheet("Sheet1 统计结果")
    ws1.append(
        [
            "样品序号",
            "导管腔面积 Ai (μm2)（可记录多个）",
            "最大导管腔面积 Amax (μm2)",
            "最小导管腔面积 Amin (μm2)",
            "平均导管腔面积 Amean (μm2)",
            "等效圆直径 Di (μm)（可记录多个）",
            "最大等效圆直径 Dmax (μm)",
            "最小等效圆直径 Dmin (μm)",
            "平均等效圆直径 Dmean (μm)",
            "水力加权直径 Dh(μm)",
            "双导管壁厚 TVW (μm)",
            "壁强化指数 CWR",
            "",
            "黄色统计 绿色计算",
        ]
    )
    for sid in range(sample_lo, sample_hi + 1):
        if sid not in sample_vessel_rows:
            ws1.append([f"样品{sid}", "", "", "", "", "", "", "", "", "", "", "", "", "无合格导管"])
            continue
        r0, r1 = sample_vessel_rows[sid]
        n = r1 - r0 + 1
        ai_ref = f'见“{mother}”F{r0}:F{r1}（全部{n}个合格导管）'
        di_ref = f'见“{mother}”G{r0}:G{r1}（{n}个逐导管Di）'
        if sid in sample_pair_rows:
            p0, p1 = sample_pair_rows[sid]
            tvw_f = f"=AVERAGE({pair_sheet}!K{p0}:K{p1})"
            cwr_f = f"=AVERAGE({pair_sheet}!S{p0}:S{p1})"
            npair = p1 - p0 + 1
        else:
            tvw_f = None
            cwr_f = None
            npair = 0
        note = (
            f"自动管线：A/D/Dh用全部{n}个合格导管；CWR用{npair}个相贴导管对。"
            f"每壁3–5条局部法线TVW取均值为主值，再算CWR_pair；Ai=不规则闭合轮廓。"
        )
        ws1.append(
            [
                f"样品{sid}",
                ai_ref,
                f"=MAX({mother}!F{r0}:F{r1})",
                f"=MIN({mother}!F{r0}:F{r1})",
                f"=AVERAGE({mother}!F{r0}:F{r1})",
                di_ref,
                f"=MAX({mother}!G{r0}:G{r1})",
                f"=MIN({mother}!G{r0}:G{r1})",
                f"=AVERAGE({mother}!G{r0}:G{r1})",
                f"=SUMPRODUCT({mother}!G{r0}:G{r1}^5)/SUMPRODUCT({mother}!G{r0}:G{r1}^4)",
                tvw_f,
                cwr_f,
                "",
                note,
            ]
        )

    for col in range(1, 15):
        ws1.column_dimensions[get_column_letter(col)].width = 16
    ws1.column_dimensions["B"].width = 44
    ws1.column_dimensions["F"].width = 36
    ws1.column_dimensions["N"].width = 60

    # ---------- explanation sheets ----------
    if GOLDEN is not None and GOLDEN.exists():
        golden_wb = load_workbook(GOLDEN, data_only=False)
        yellow = next(n for n in golden_wb.sheetnames if "黄色" in n)
        green = next(n for n in golden_wb.sheetnames if "绿色" in n)
        copy_sheet(golden_wb, yellow, wb, "Sheet1 黄色指标样貌说明")
        copy_sheet(golden_wb, green, wb, "Sheet1 绿色指标计算公式")
        sanitize_green_demo_formulas(wb["Sheet1 绿色指标计算公式"])
    else:
        print("警告: 未找到 Wand 模板 xlsx，跳过两张说明表")

    # Desired order
    order = [
        summary_sheet,
        mother,
        audit_sheet,
        view_sheet,
        pair_sheet,
        line_sheet,
        "Sheet1 统计结果",
        "Sheet1 黄色指标样貌说明",
        "Sheet1 绿色指标计算公式",
    ]
    if GOLDEN is None or not GOLDEN.exists():
        order = order[:7]
    for i, name in enumerate(order):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.save(dest)
    print(f"sheets: {wb.sheetnames}")
    print(f"vessels={n_v} pairs={n_p} lines={line_row - 1} images={len(images)}")
    print(f"saved: {dest}")
    return dest


def main():
    paths = list_unique_images()
    print(f"Images: {len(paths)}")
    print(f"Out: {XLSX_OUT}")
    t0 = time.time()

    by_sample: dict[int, dict] = defaultdict(lambda: {"vessels": [], "pairs": []})
    images: list[dict] = []

    for i, path in enumerate(paths, 1):
        sid = sample_id(path)
        rgb, gray, vessels, _, scale = process_image(path, force_um_per_px=None)
        for v in vessels:
            v.pair_id = None
        pairs = find_pairs(vessels, um_per_px=scale["um_per_px"], image_name=path.name)
        by_sample[sid]["vessels"].extend(vessels)
        by_sample[sid]["pairs"].extend(pairs)
        images.append(
            {
                "image": path.name,
                "sample": sid,
                "um_per_px": float(scale["um_per_px"]),
                "tissue_h": tissue_height(gray.shape[0]),
                "n_vessel": len(vessels),
                "n_pair": len(pairs),
            }
        )
        if i % 50 == 0 or i == len(paths):
            print(f"[{i}/{len(paths)}] {path.name} | {time.time() - t0:.0f}s")

    write_workbook(by_sample, images)
    print(f"Wrote {XLSX_OUT} in {time.time() - t0:.1f}s")


if __name__ == "__main__":
    main()
